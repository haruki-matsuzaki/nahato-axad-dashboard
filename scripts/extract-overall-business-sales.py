#!/usr/bin/env python3
import argparse
import calendar
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX


SHEET_NAME = "◆全体売上表"
START_ROW = 57
TEXT_COLUMNS_METRIC_J = [
    (2, "unit", "Unit", 150),
    (3, "media", "媒体", 76),
    (4, "team", "チーム", 96),
    (5, "owner", "担当", 96),
    (6, "detail", "案件 / 詳細", 260),
    (8, "officialName", "案件 / 正式名称", 180),
    (9, "client", "クライアント", 180),
    (10, "metric", "項目", 76),
    (11, "total", "合計", 108),
]
TEXT_COLUMNS_METRIC_K = [
    (2, "unit", "Unit", 150),
    (3, "media", "媒体", 76),
    (4, "team", "チーム", 96),
    (5, "owner", "担当", 96),
    (6, "detail", "案件 / 詳細", 260),
    (8, "officialName", "案件 / 正式名称", 180),
    (9, "aggregationKey", "集計キー", 220),
    (10, "client", "クライアント", 180),
    (11, "metric", "項目", 76),
    (12, "total", "合計", 108),
]
METRICS = {"売上", "粗利", "利鞘", "件数"}


def main():
    parser = argparse.ArgumentParser(description="Extract rows 57+ from the overall sales worksheet.")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--month", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--sheet", default=SHEET_NAME)
    args = parser.parse_args()

    workbook = load_workbook(args.xlsx, read_only=False, data_only=True)
    if args.sheet not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {args.sheet}")

    worksheet = workbook[args.sheet]
    year, month = parse_month(args.month)
    days = calendar.monthrange(year, month)[1]
    layout = detect_layout(worksheet)
    date_start_column = layout["total_column"] + 1
    end_column = layout["total_column"] + days

    columns = []
    for column_index, key, label, width in layout["text_columns"]:
        columns.append({"index": column_index, "key": key, "label": label, "width": width, "type": "text"})
    for day in range(1, days + 1):
        current = date(year, month, day)
        columns.append(
            {
                "index": date_start_column + day - 1,
                "key": current.isoformat(),
                "label": f"{month}/{day}",
                "weekday": "月火水木金土日"[current.weekday()],
                "width": 104,
                "type": "value",
            }
        )

    rows = []
    start_row = START_ROW + 1 if layout["has_header"] else START_ROW
    last_row = start_row - 1
    for row_offset, row_cells in enumerate(
        worksheet.iter_rows(min_row=start_row, min_col=1, max_col=end_column),
        start=start_row,
    ):
        values = []
        styles = []
        has_content = False
        for column in columns:
            cell = row_cells[column["index"] - 1] if column["index"] - 1 < len(row_cells) else None
            raw_value = cell.value if cell else None
            value = normalize_value(raw_value)
            values.append(value)
            background = cell_background(cell) if cell and 2 <= column["index"] <= 10 else None
            styles.append({"background": background} if background else "")
            if value not in (None, ""):
                has_content = True
        if not has_content:
            continue
        last_row = row_offset
        row_payload = {"index": row_offset, "values": values}
        if any(styles):
            row_payload["styles"] = styles
        rows.append(row_payload)

    payload = {
        "month": args.month,
        "source": {
            "workbook": Path(args.xlsx).name,
            "sheetName": args.sheet,
            "range": f"{start_row}:{last_row}",
            "generatedAt": datetime.now().isoformat(),
        },
        "columns": columns,
        "rows": rows,
    }

    output_path = Path(args.out)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def detect_layout(worksheet):
    row_values = {column: normalize_text(worksheet.cell(START_ROW, column).value) for column in range(2, 13)}
    has_header = row_values.get(2) in {"Unit", "事業部", "部署"} or row_values.get(11) == "合計"

    if has_header or row_values.get(10) in METRICS:
        return {
            "has_header": has_header,
            "metric_column": 10,
            "total_column": 11,
            "text_columns": TEXT_COLUMNS_METRIC_J,
        }

    return {
        "has_header": False,
        "metric_column": 11,
        "total_column": 12,
        "text_columns": TEXT_COLUMNS_METRIC_K,
    }


def parse_month(month_id):
    year, month = month_id.split("-")
    return int(year), int(month)


def normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_text(value):
    return str(value or "").strip()


def cell_background(cell):
    if not cell or not cell.fill or cell.fill.patternType != "solid":
        return None
    return color_to_css(cell.fill.fgColor)


def color_to_css(color):
    if not color:
        return None
    if color.type == "rgb" and color.rgb:
        value = str(color.rgb)
        if value == "00000000":
            return None
        if len(value) == 8:
            value = value[2:]
        return f"#{value.lower()}"
    if color.type == "indexed" and color.indexed is not None:
        try:
            value = COLOR_INDEX[color.indexed]
        except IndexError:
            return None
        if len(value) == 8:
            value = value[2:]
        return f"#{value.lower()}"
    return None


if __name__ == "__main__":
    main()
