#!/usr/bin/env python3
import argparse
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def main():
    parser = argparse.ArgumentParser(description="Extract worksheet values from an xlsx file as JSON rows.")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    workbook = load_workbook(args.xlsx, read_only=False, data_only=True)
    if args.sheet not in workbook.sheetnames:
      raise SystemExit(f"Sheet not found: {args.sheet}")

    worksheet = workbook[args.sheet]
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        values = [normalize_value(cell) for cell in row]
        while values and values[-1] is None:
            values.pop()
        rows.append(values)

    while rows and not rows[-1]:
        rows.pop()

    output_path = Path(args.out)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")


def normalize_value(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


if __name__ == "__main__":
    main()
