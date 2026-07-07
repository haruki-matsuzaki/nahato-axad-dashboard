#!/usr/bin/env python3
import argparse
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter


DEFAULT_BORDER = "#d9d9d9"


def main():
    parser = argparse.ArgumentParser(description="Extract a formatted worksheet range for the overall sales view.")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--month", required=True)
    parser.add_argument("--start-row", type=int, required=True)
    parser.add_argument("--end-row", type=int, required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    workbook = load_workbook(args.xlsx, read_only=False, data_only=True)
    if args.sheet not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {args.sheet}")

    worksheet = workbook[args.sheet]
    visible_columns = [
        column
        for column in range(1, worksheet.max_column + 1)
        if not worksheet.column_dimensions[get_column_letter(column)].hidden
    ]
    visible_column_set = set(visible_columns)

    merges = collect_merges(worksheet, args.start_row, args.end_row, visible_column_set)
    skipped = set()
    for merge in merges:
        for row in range(merge["row"], merge["row"] + merge["rowspan"]):
            for column in merge["visibleColumns"]:
                if row == merge["row"] and column == merge["column"]:
                    continue
                skipped.add((row, column))

    rows = []
    for row_index in range(args.start_row, args.end_row + 1):
        row_dimension = worksheet.row_dimensions[row_index]
        row = {
            "index": row_index,
            "height": row_height_to_px(row_dimension.height),
            "hidden": bool(row_dimension.hidden),
            "cells": [],
        }
        for column_index in visible_columns:
            if (row_index, column_index) in skipped:
                row["cells"].append({"skip": True})
                continue

            cell = worksheet.cell(row_index, column_index)
            merge = merge_for_cell(merges, row_index, column_index)
            row["cells"].append(
                {
                    "address": f"{get_column_letter(column_index)}{row_index}",
                    "rowspan": merge["rowspan"] if merge else 1,
                    "colspan": len(merge["visibleColumns"]) if merge else 1,
                    "value": normalize_value(cell.value),
                    "text": format_display_value(cell.value, cell.number_format),
                    "style": cell_style(cell),
                }
            )
        rows.append(row)

    columns = []
    for column_index in visible_columns:
        column_letter = get_column_letter(column_index)
        dimension = worksheet.column_dimensions[column_letter]
        columns.append(
            {
                "index": column_index,
                "letter": column_letter,
                "width": column_width_to_px(dimension.width),
            }
        )

    payload = {
        "month": args.month,
        "source": {
            "workbook": Path(args.xlsx).name,
            "sheetName": args.sheet,
            "range": f"{args.start_row}:{args.end_row}",
            "generatedAt": datetime.now().isoformat(),
        },
        "columns": columns,
        "rows": rows,
    }

    output_path = Path(args.out)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def collect_merges(worksheet, start_row, end_row, visible_column_set):
    merges = []
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_row < start_row or merged_range.min_row > end_row:
            continue
        visible_columns = [
            column
            for column in range(merged_range.min_col, merged_range.max_col + 1)
            if column in visible_column_set
        ]
        if not visible_columns:
            continue
        row = max(merged_range.min_row, start_row)
        rowspan = min(merged_range.max_row, end_row) - row + 1
        merges.append(
            {
                "row": row,
                "column": visible_columns[0],
                "rowspan": rowspan,
                "visibleColumns": visible_columns,
            }
        )
    return merges


def merge_for_cell(merges, row, column):
    for merge in merges:
        if merge["row"] == row and merge["column"] == column:
            return merge
    return None


def normalize_value(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


def format_display_value(value, number_format):
    if value is None:
        return ""
    if isinstance(value, datetime):
        if "m" in number_format.lower() and "d" in number_format.lower():
            return f"{value.month}/{value.day}"
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        if "m" in number_format.lower() and "d" in number_format.lower():
            return f"{value.month}/{value.day}"
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        number = float(value)
        if number_format == "0.00%":
            return f"{number * 100:.2f}%"
        if "¥" in number_format or "[$¥" in number_format:
            return f"¥{round(number):,}"
        if "#,##0" in number_format:
            return f"{round(number):,}"
        if number.is_integer():
            return str(int(number))
        return f"{number:.10g}"
    return str(value)


def cell_style(cell):
    styles = []
    fill_color = color_to_css(cell.fill.fgColor)
    if cell.fill.patternType == "solid" and fill_color:
        styles.append(f"background-color:{fill_color}")

    font = cell.font
    if font.name:
        styles.append(f"font-family:{css_string(font.name)}")
    if font.sz:
        styles.append(f"font-size:{font.sz}pt")
    if font.bold:
        styles.append("font-weight:700")
    if font.italic:
        styles.append("font-style:italic")
    if font.underline:
        styles.append("text-decoration:underline")
    font_color = color_to_css(font.color)
    if font_color:
        styles.append(f"color:{font_color}")

    alignment = cell.alignment
    if alignment.horizontal:
        styles.append(f"text-align:{horizontal_alignment(alignment.horizontal)}")
    if alignment.vertical:
        styles.append(f"vertical-align:{vertical_alignment(alignment.vertical)}")
    if alignment.wrap_text:
        styles.append("white-space:pre-wrap")

    for side_name in ["left", "right", "top", "bottom"]:
        side = getattr(cell.border, side_name)
        border = border_to_css(side)
        if border:
            styles.append(f"border-{side_name}:{border}")

    return ";".join(styles)


def color_to_css(color):
    if not color:
        return None
    if color.type == "rgb" and color.rgb:
        value = str(color.rgb)
        if value == "00000000":
            return None
        if len(value) == 8:
            value = value[2:]
        return f"#{value.lower()}"
    if color.type == "indexed" and color.indexed is not None:
        try:
            value = COLOR_INDEX[color.indexed]
        except IndexError:
            return None
        if len(value) == 8:
            value = value[2:]
        return f"#{value.lower()}"
    return None


def border_to_css(side):
    if not side or not side.style:
        return None
    width = {
        "hair": "1px",
        "thin": "1px",
        "medium": "2px",
        "thick": "3px",
        "double": "3px",
    }.get(side.style, "1px")
    line_style = "double" if side.style == "double" else "solid"
    color = color_to_css(side.color) or DEFAULT_BORDER
    return f"{width} {line_style} {color}"


def horizontal_alignment(value):
    return {
        "centerContinuous": "center",
        "distributed": "center",
        "fill": "left",
        "general": "left",
    }.get(value, value)


def vertical_alignment(value):
    return {
        "center": "middle",
    }.get(value, value)


def column_width_to_px(width):
    if not width:
        return 64
    return max(0, round(width * 7 + 5))


def row_height_to_px(height):
    if not height:
        return None
    return round(height * 1.333)


def css_string(value):
    return "'" + str(value).replace("\\", "\\\\").replace("'", "\\'") + "'"


if __name__ == "__main__":
    main()
